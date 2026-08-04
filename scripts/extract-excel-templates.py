#!/usr/bin/env python3
"""
extract-excel-templates.py — mine a corpus of reference menu spreadsheets into
knowledge/excel-template-profiles.json (the "Template Knowledge Base").

Why this exists
---------------
The Excel exporter must not hard-code the look of one sample file. Instead it
renders from a *design system* — and that design system is DERIVED from the
reference corpus by this script, exactly the way scripts/build-knowledge-base.py
derives the RAG corpus from the source PDFs.

Re-run it whenever new reference files are added:

    pip install -r scripts/requirements.txt
    python3 scripts/extract-excel-templates.py --src ./final_sample

What it extracts per file
-------------------------
  • layout grammar   — the ordered list of block roles (TITLE / META / TABLE /
                       SECTION / NOTE / SPACER) inferred from cell styling
  • style mapping    — fill + font + alignment for every role
  • geometry         — column widths, row heights, merge spans
  • page setup       — orientation, margins, print area, freeze panes

and then aggregates the whole corpus into a single canonical theme (the modal
value of every token) plus the observed variants.

NOTE: exceljs cannot READ workbooks written by openpyxl/LibreOffice (they omit
docProps/app.xml, which exceljs's reader requires), which is why extraction is
done here in Python with openpyxl rather than in Node.
"""
import argparse
import collections
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Thiếu openpyxl — chạy: pip install -r scripts/requirements.txt")

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = os.path.join(HERE, "..", "knowledge", "excel-template-profiles.json")


# ───────────────────────────── helpers ─────────────────────────────

def rgb(color):
    """openpyxl Color -> 'FFRRGGBB' or None (theme/indexed colors are skipped)."""
    if color is None:
        return None
    try:
        if color.type == "rgb" and color.rgb and color.rgb != "00000000":
            return color.rgb
    except Exception:
        pass
    return None


def fill_of(cell):
    if cell.fill is None or not cell.fill.patternType:
        return None
    return rgb(cell.fill.fgColor)


def row_is_empty(ws, r):
    return all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ws.max_column + 1))


def row_merge_span(ws, r):
    """Return the merge width if row r starts with a full-width merge, else 0."""
    for m in ws.merged_cells.ranges:
        if m.min_row == r and m.max_row == r and m.min_col == 1:
            return m.max_col
    return 0


def classify_row(ws, r):
    """Infer the semantic role of a row from its styling — the layout grammar."""
    cell = ws.cell(row=r, column=1)
    fill = fill_of(cell)
    font = cell.font
    merged = row_merge_span(ws, r) > 0

    if row_is_empty(ws, r) and not fill:
        return "spacer"
    if merged and font.sz and font.sz >= 14:
        return "title"
    if merged and font.i:
        return "note"
    if merged and font.b:
        return "sectionHeader"
    if fill and font.b and rgb(font.color) == "FFFFFFFF":
        return "tableHeader"
    if fill and font.i:
        return "meta"
    if fill and font.b:
        return "rowLabel"
    if fill:
        return "rowLabel"
    return "dataRow"


def style_of(cell):
    return {
        "fill": fill_of(cell),
        "fontColor": rgb(cell.font.color),
        "bold": bool(cell.font.b),
        "italic": bool(cell.font.i),
        "size": cell.font.sz,
        "name": cell.font.name,
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
        "wrap": bool(cell.alignment.wrap_text),
    }


def profile_sheet(ws):
    """Full structural + stylistic fingerprint of one worksheet."""
    roles = [classify_row(ws, r) for r in range(1, ws.max_row + 1)]

    # style sample per role (first occurrence wins — they are uniform in practice)
    role_styles = {}
    for idx, role in enumerate(roles, start=1):
        if role in role_styles or role == "spacer":
            continue
        role_styles[role] = style_of(ws.cell(row=idx, column=1))

    # the body style is the 2nd column of the first data-bearing row under a header
    body_style = None
    for idx, role in enumerate(roles, start=1):
        if role == "rowLabel" and ws.max_column > 1:
            body_style = style_of(ws.cell(row=idx, column=2))
            break

    # numeric columns: trailing columns whose data cells are centred + tinted
    numeric_fill = None
    for idx, role in enumerate(roles, start=1):
        if role != "rowLabel":
            continue
        for c in range(ws.max_column, 1, -1):
            cell = ws.cell(row=idx, column=c)
            if cell.alignment.horizontal == "center" and fill_of(cell):
                numeric_fill = fill_of(cell)
        break

    heights = collections.Counter()
    for k, rd in ws.row_dimensions.items():
        if rd.height:
            heights[rd.height] += 1

    return {
        "sheetName": ws.title,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "grammar": compress(roles),
        "roleStyles": role_styles,
        "bodyStyle": body_style,
        "numericFill": numeric_fill,
        "columnWidths": {k: cd.width for k, cd in sorted(ws.column_dimensions.items()) if cd.width},
        "rowHeights": dict(heights.most_common(6)),
        "merges": [str(m) for m in ws.merged_cells.ranges],
        "freezePanes": ws.freeze_panes,
        "pageSetup": {
            "orientation": ws.page_setup.orientation,
            "paperSize": ws.page_setup.paperSize,
            "fitToWidth": ws.page_setup.fitToWidth,
            "printArea": ws.print_area or None,
            "margins": {
                "left": ws.page_margins.left,
                "right": ws.page_margins.right,
                "top": ws.page_margins.top,
                "bottom": ws.page_margins.bottom,
            },
        },
        "hasBorders": any(
            c.border and c.border.left and c.border.left.style
            for row in ws.iter_rows() for c in row
        ),
        "images": len(ws._images),
    }


def compress(roles):
    """['title','meta','meta','spacer',...] -> [{'role':'meta','count':2}, ...]"""
    out = []
    for role in roles:
        if out and out[-1]["role"] == role:
            out[-1]["count"] += 1
        else:
            out.append({"role": role, "count": 1})
    return out


def modal(counter, default=None):
    return counter.most_common(1)[0][0] if counter else default


# ───────────────────────────── main ─────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="Thư mục chứa các file .xlsx/.xlsm mẫu")
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    files = sorted(
        f for f in os.listdir(args.src)
        if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
    )
    if not files:
        sys.exit(f"Không tìm thấy file Excel nào trong {args.src}")

    profiles = []
    tokens = collections.defaultdict(collections.Counter)

    for name in files:
        path = os.path.join(args.src, name)
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:  # noqa: BLE001 — a bad sample must not kill the run
            print(f"  ⚠️  bỏ qua {name}: {e}")
            continue

        ws = wb.worksheets[0]
        prof = profile_sheet(ws)
        # An index/link sheet (no title block) is not a menu template.
        if not any(b["role"] == "title" for b in prof["grammar"]):
            print(f"  ↷  bỏ qua {name} (không phải template thực đơn)")
            continue

        prof["file"] = name
        profiles.append(prof)

        for role, st in prof["roleStyles"].items():
            if st.get("fill"):
                tokens[f"{role}.fill"][st["fill"]] += 1
            if st.get("fontColor"):
                tokens[f"{role}.fontColor"][st["fontColor"]] += 1
            if st.get("size"):
                tokens[f"{role}.size"][st["size"]] += 1
            if st.get("name"):
                tokens["font.name"][st["name"]] += 1
        if prof.get("numericFill"):
            tokens["numeric.fill"][prof["numericFill"]] += 1
        for w in prof["columnWidths"].values():
            tokens["columnWidth"][round(w)] += 1

    canonical = {k: modal(v) for k, v in tokens.items()}
    variants = {k: v.most_common(6) for k, v in tokens.items()}

    out = {
        "_generatedBy": "scripts/extract-excel-templates.py",
        "_corpusSize": len(profiles),
        "_note": (
            "Design system suy ra từ corpus file mẫu — KHÔNG hard-code theo một file. "
            "lib/excel/theme.js đọc `canonical` làm theme mặc định."
        ),
        "canonical": canonical,
        "variants": variants,
        "grammar": most_common_grammar(profiles),
        "profiles": profiles,
    }

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=2)

    # Compact runtime artifact — lib/excel/theme.js imports THIS, not the full
    # corpus, so the serverless bundle stays small.
    theme_path = os.path.join(os.path.dirname(os.path.abspath(args.out)), "excel-theme.json")
    with open(theme_path, "w", encoding="utf-8") as fh:
        json.dump(
            {
                "_generatedBy": "scripts/extract-excel-templates.py",
                "_corpusSize": len(profiles),
                "tokens": canonical,
                "variants": {k: [c for c, _ in v] for k, v in variants.items()},
            },
            fh,
            ensure_ascii=False,
            indent=2,
        )

    print(f"✅ {len(profiles)} template → {args.out}")
    print(f"✅ theme rút gọn      → {theme_path}")
    print(f"   Canonical tokens: {json.dumps(canonical, ensure_ascii=False)[:200]}...")


def most_common_grammar(profiles):
    """The most frequent block sequence — the 'shape' a generated file should have."""
    seqs = collections.Counter(
        tuple(b["role"] for b in p["grammar"] if b["role"] != "spacer") for p in profiles
    )
    return [{"sequence": list(seq), "count": n} for seq, n in seqs.most_common(8)]


if __name__ == "__main__":
    main()
